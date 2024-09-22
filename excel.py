#Created by Eric Buchanan
# Objective: Open Excel file, take data from two specific sheets and compare that data. Once a match is found take address for matching query and assign it to persons address.
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border, colors

#Load Workbook into wb
wb = load_workbook('E:\VSCode Workspace\document.xlsx')

matchFirst = []
matchLast = []
practiceName = []
practiceAddress = []
practiceExtra = []
foundAddress = []
tempFirst = ''
tempLast = ''

#declares and assigns variables for indivdual sheets
matchData = wb['Notification']
practiceWB = wb['Sheet1']

#For loops to take data from each sheet and row and append array for every new variable. 
for row in matchData.iter_rows(min_row = 2,min_col = 3,max_row = 12997, max_col = 3):
        for cell in row:
            matchFirst.append(cell.value)

greenFill = PatternFill(start_color = '008000', end_color = '008000', fill_type='solid')

for row in matchData.iter_rows(min_row = 2,min_col = 5,max_row = 12997, max_col = 5):
        for cell in row:
            matchLast.append(cell.value)

for row in practiceWB.iter_rows(min_row = 2,min_col = 3,max_row = 12736, max_col = 3):
        for cell in row:
            if cell.value == None:
                practiceName.append("fhueabfghebahgi")
            else:
                practiceName.append(cell.value)      

for row in practiceWB.iter_rows(min_row = 2,min_col = 6,max_row = 12736, max_col = 6):
        for cell in row:
            if cell.value == None or cell.value == ',' or '@' in cell.value or 'email' in cell.value: 
                practiceAddress.append("not found")  
            else:
                practiceAddress.append(cell.value) 

for row in practiceWB.iter_rows(min_row = 2,min_col = 14,max_row = 12736, max_col = 14):
        for cell in row:
            if cell.value == None: 
                practiceExtra.append("***")  
            else:
                practiceExtra.append(cell.value)     

num = 0
for z in practiceAddress:
    if practiceAddress[num] == "not found":
        for row in practiceWB.iter_rows(min_row = num,min_col = 9,max_row = num, max_col = 9):
            if not cell.value == ',,' or '@' in cell.value or 'email' in cell.value:
                practiceAddress[num] = cell.value
            else:
                for row in practiceWB.iter_rows(min_row = num,max_row = num,min_col = 10,max_col = 10):
                    if not cell.value == ',,' or '@' in cell.value or 'email' in cell.value:
                        practiceAddress[num] = cell.value
    num += 1        

#Ensure lengths stay constant
matchLength = len(matchFirst)
practiceLength = len(practiceName)

#print (matchLength, practiceLength)

# For loop increment i till it equals matchLength set tempFirst and tempLast
for i in range(matchLength):
    tempFirst = matchFirst[i].lower()
    tempLast = matchLast[i].lower()
    tempAddress = 'notfound'
    for j in range(practiceLength):
        tempName = practiceName[j].lower()
        if (tempFirst in tempName and tempLast in tempName):        
            tempAddress = practiceAddress[j]
            foundAddress.append(practiceAddress[j])     
            break 
        else:
            continue
    if tempAddress == 'notfound':   # if tempAddress is equal to not found then put in foundAddress list as Not Found
        foundAddress.append("Not found")

#print (practiceExtra)

wb.active = matchData
ws = wb.active
#for Loop to go down the foundAddress list and writes it to column24
for i, address in enumerate(foundAddress, start=2):
    ws.cell(row=i, column=24, value=address)
count = 2  # Start from the second row since headers are in the first row
ss = 0
win = 0
lose = 0

for idx, row in enumerate(matchData.iter_rows(min_row=2, min_col=16, max_row=12997, max_col=16), start=2):
    cell = row[0]  # Since we're only dealing with one column
    if cell.value is None:
        continue
    else:
        tempValue = str(cell.value).replace('-', '').replace(' ', '')
        match_found = False  # Flag to indicate if a match was found
        
        for k in range(practiceLength):
            practiceTempValue = str(practiceExtra[k]).replace('-', '').replace(' ', '')
            
            if tempValue == practiceTempValue:
                ss += 1
                cell.fill = greenFill  # Fill the cell with green color
                
                # Clean the addresses for comparison
                found_address_clean = str(foundAddress[idx - 2]).replace(' ', '').replace(',', '').lower()
                ws_address_clean = str(ws.cell(row=idx, column=24).value).replace(' ', '').replace(',', '').lower()
                
                if found_address_clean == ws_address_clean:
                    print("Worked", idx, found_address_clean, "|", ws_address_clean)
                    if ws.cell(row=idx, column=24).value != "Not found":
                        ws.cell(row=idx, column=24).fill = greenFill
                    win += 1
                else:
                    print("FAILED", idx, found_address_clean, "|", ws_address_clean)
                    lose += 1
                match_found = True
                break  # Exit the inner loop since we found a match
        
        if not match_found:
            # Handle the case where no match was found, if necessary
            pass

print("Win:", win, " Lose:", lose)
print("Total Matches Found:", ss)



#debug statements
#print (len(foundAddress),matchLength,practiceLength)    
#print(foundAddress)

#save file whenever we get to that       
wb.save('E:\VSCode Workspace\ balances.xlsx')